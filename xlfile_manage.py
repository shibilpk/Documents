# install package
    # pip install openpyxl
# put the xl file to app that need to assess the data
import openpyxl as xl

wb = xl.load_workbook('xlfile_name.xlsl')
sheet = wb['Sheet1'] #specify the sheet
#cell = sheet['a1'] #a1 cell
#cell = sheet.cell(1 ,1) # a1 cell
#sheet.max_row # to get the maximum length of row

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    # print(cell.value) # to print value
    corrented_price = cell.value * .7
    corrented_price_cell = sheet.cell(row, 4)
    corrented_price_cell.value = corrented_price

wb.save('name.xlsx')
